from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, numbers
import time, re
from keuntungan import keuntungan
from openpyxl.formula.translate import Translator
from open_value_excel import get_duit
cell_F = None
alpha_harga = None
number_harga = None
def debit(namaprdk,jumlahprdk,hrgprdk,produk_belanja,jumlah_belanja,harga_belanja):
    date = datetime.now()
    yellow = "00FFFF00"
    green = "000FF000"
    red = "FFFF0000"
    color = "000FFF00"
    thin = Side(border_style="thin",color="00000000")
    try:
        wb = load_workbook("debit.xlsx")
        ws = wb.active
        # ws = wb[str(date.strftime("%B"))]
        if str(date.strftime("%B")) in wb.sheetnames:
            ws = wb[str(date.strftime("%B"))]
        else: 
            # pass
            ws = wb.create_sheet(str(date.strftime("%B"))) 
            ws = wb[str(date.strftime("%B"))]
            ws.title = date.strftime("%B")
            wb.save("debit.xlsx")

        """==KAS KELUAR=="""
        ws.column_dimensions['H'].auto_size = True
        ws.column_dimensions['I'].auto_size = True
        ws.column_dimensions['J'].auto_size = True
        ws.cell(row=1, column=8).value = 'Belanja Produk'
        ws.cell(row=1, column=8).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=8).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        ws.cell(row=1, column=9).value = 'Jumlah Belanja'
        ws.cell(row=1, column=9).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=9).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        ws.cell(row=1, column=10).value = 'Harga Belanja'
        ws.cell(row=1, column=10).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=10).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        """=============="""
        ws.column_dimensions['A'].auto_size = True
        ws.column_dimensions['B'].auto_size = True
        ws.column_dimensions['C'].auto_size = True
        ws.column_dimensions['D'].auto_size = True
        ws.column_dimensions['E'].auto_size = True
        ws.column_dimensions['F'].auto_size = True
        ws.cell(row=1, column=1).value = 'Tanggal'
        ws.cell(row=1, column=1).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=1).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=2).value = 'Waktu'
        ws.cell(row=1, column=2).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=2).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=3).value = 'Nama Produk'
        ws.cell(row=1, column=3).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=3).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=4).value = 'Jumlah Produk'
        ws.cell(row=1, column=4).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=4).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=5).value = 'Harga Produk'
        ws.cell(row=1, column=5).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=5).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        wb.save("debit.xlsx")
    except:
        wb = Workbook()
        if str(date.strftime("%B")) in wb.sheetnames:
            ws = wb[str(date.strftime("%B"))]
        else: ws = wb.active
        
        ws.title = date.strftime("%B")
        
        ws.column_dimensions['A'].auto_size = True
        ws.column_dimensions['B'].auto_size = True
        ws.column_dimensions['C'].auto_size = True
        ws.column_dimensions['D'].auto_size = True
        ws.column_dimensions['E'].auto_size = True
        ws.column_dimensions['F'].auto_size = True
        ws.column_dimensions['G'].auto_size = True
        ws.cell(row=1, column=1).value = 'Tanggal'
        ws.cell(row=1, column=1).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=1).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=2).value = 'Waktu'
        ws.cell(row=1, column=2).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=2).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=3).value = 'Nama Produk'
        ws.cell(row=1, column=3).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=3).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=4).value = 'Jumlah Produk'
        ws.cell(row=1, column=4).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=4).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws.cell(row=1, column=5).value = 'Harga Produk'
        ws.cell(row=1, column=5).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=5).fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        """==KAS KELUAR=="""
        ws.column_dimensions['H'].auto_size = True
        ws.column_dimensions['I'].auto_size = True
        ws.column_dimensions['J'].auto_size = True
        ws.cell(row=1, column=8).value = 'Belanja Produk'
        ws.cell(row=1, column=8).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=8).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        ws.cell(row=1, column=9).value = 'Jumlah Belanjaan'
        ws.cell(row=1, column=9).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=9).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        ws.cell(row=1, column=10).value = 'Harga Belanja'
        ws.cell(row=1, column=10).border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.cell(row=1, column=10).fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        """=============="""
        wb.save("debit.xlsx")
    

    col_A = ws["A" + str(len(ws['A']))]
    current_date = date.strftime("%D")
    current_time = date.strftime("%H:%M:%S")
    ws.append([current_date,current_time,namaprdk,jumlahprdk,int(hrgprdk)])
    wb.save('debit.xlsx')
    cellA = ws['A' + str(len(ws['A']))]
    cellB = ws['B' + str(len(ws['B']))]
    cellC = ws['C' + str(len(ws['C']))]
    cellD = ws['D' + str(len(ws['D']))]
    cellE = ws['E' + str(len(ws['E']))]
    cellF = ws['F' + str(len(ws['F']))]
    cellH = ws['H' + str(len(ws['H']))]
    cellI = ws['I' + str(len(ws['I']))]
    cellJ = ws['J' + str(len(ws['J']))]
    cellA.number_format = 'dd/mm/yyyy'

    cellH.value = produk_belanja
    cellI.value = jumlah_belanja
    cellJ.value = harga_belanja
    
    target = 'Tanggal'
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                # print("Found at cell:", cell.coordinate)
                tgl = cell.coordinate
                # print(tgl)
    regex = re.compile(r'(\d+|\s+)')
    tgl = regex.split(tgl)
    alpha = tgl[0]
    num = int(tgl[1]) + 1
    cell_tgl = ws[alpha + str(num)]


    if cell_tgl.value != current_date:
        cellA.value = 'Tanggal'
        cell_tanggal = ws['A' + str(len(ws['A'])+1)]
        cell_tanggal.value = current_date
        cellB.value = 'Waktu'
        cell_waktu = ws['B' + str(len(ws['B']))]
        cell_waktu.value = current_time
        cellC.value = 'Nama Produk'
        cell_nama = ws['C' + str(len(ws['C']))]
        cell_nama.value = namaprdk
        cellD.value = 'Jumlah Produk'
        cell_jmlh = ws['D' + str(len(ws['D']))]
        cell_jmlh.value = jumlahprdk
        cellE.value = 'Harga Produk'
        cell_harga = ws['E' + str(len(ws['E']))]
        cell_harga.value = int(hrgprdk)
        cell_harga.number_format = 'Rp #,##0.00'
        cellH.value = "Belanja Produk"
        cell_belanja_produk = ws['H' + str(len(ws['H']))]
        cell_belanja_produk.value = produk_belanja
        cellI.value = "Jumlah Belanjaan"
        cell_jumlah_belanjaan = ws['I' + str(len(ws['I']))]
        cell_jumlah_belanjaan.value = jumlah_belanja
        cellJ.value = "Harga Belanja"
        cell_harga_belanjaan = ws['J' + str(len(ws['J']))]
        cell_harga_belanjaan.value = harga_belanja
        cell_belanja_produk.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_harga_belanjaan.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_jumlah_belanjaan.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellA.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellB.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellC.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellD.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellE.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellH.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellI.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellJ.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cellA.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        cellB.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        cellC.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        cellD.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        cellE.fill =PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        cellH.fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        cellI.fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        cellJ.fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        
        
    else: pass
    
    # for row in ws.iter_rows(min_col=4,max_col=4,):
    #     print(row)

    cellA.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellB.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellC.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellD.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellE.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellH.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellI.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellJ.border = Border(top=thin,left=thin,right=thin,bottom=thin)
    cellE.number_format = 'Rp #,##0.00'
    cellJ.number_format = 'Rp #,##0.00'
    
    wb.save('debit.xlsx')
    target = 'Harga Produk'
    global list_cell
    list_cell = []
    list_cellJ = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                hrg = cell.coordinate
                hrg = regex.split(hrg)
                alpha_hrg = hrg[0]
                alpha_harga = alpha_hrg
                num_hrg = int(hrg[1]) + 1
                number_harga = num_hrg
                cell_hrg = ws[alpha_hrg + str(num_hrg)]
                F = ws['F' + str(num_hrg)]
                G = ws['G' + str(num_hrg)]
                K = ws['K' + str(num_hrg)]
                cell_omset = ws['F' + str(num_hrg - 1)]
                cell_pengeluaran = ws['G' + str(num_hrg - 1)]
                cell_lababersih =  ws['K' + str(num_hrg - 1)]
                cell_keuntungan = ws['F' + str(num_hrg + 1)]
                cell_hasil = ws['F' + str(num_hrg + 2)]
                cell_F = F
                cell_G = G
                cell_K = K
                for row in ws.iter_rows(min_col=5,max_col=5,min_row=int(hrg[1])):
                    for cell in row:
                        if cell.value == target:
                            break
                        break
                    break
                
                list_cell.append(cell.coordinate)
                for row in ws.iter_rows(min_col=10,max_col=10,min_row=int(hrg[1])):
                    for cell in row:
                        if cell.value == "Harga Belanja":
                            break
                        break
                    break
                list_cellJ.append(cell.coordinate)
                
    if len(list_cell) == len(list_cell): 
        avg = keuntungan()
        cell_keuntungan.value = f"Keuntungan({avg})="
        cell_omset.value = "Omset Harian = "
        cell_pengeluaran.value = "Pengeluaran Harian ="
        cell_lababersih.value = "selisih harian ="
        cell_omset.fill = PatternFill(start_color=green, end_color=green,
                                        fill_type = "solid")
        cell_pengeluaran.fill = PatternFill(start_color=green, end_color=green,
                                        fill_type = "solid")
        cell_lababersih.fill = PatternFill(start_color=green, end_color=green,
                                        fill_type = "solid")
        ws['M1'].fill = PatternFill(start_color=green, end_color=green,
                                        fill_type = "solid")
        ws['M3'].fill = PatternFill(start_color=red, end_color=red,
                                        fill_type = "solid")
        ws['M5'].fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
        ws['M1'].border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws['M3'].border = Border(top=thin,left=thin,right=thin,bottom=thin)
        ws['M5'].border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_pengeluaran.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_omset.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_lababersih.border = Border(top=thin,left=thin,right=thin,bottom=thin)
        cell_F.value = f"=SUM({list_cell[-1]}:{'E' + str(len(ws['E']))})" 
        cell_G.value = f"=SUM({list_cellJ[-1]}:{'J' + str(len(ws['J']))})" 
        cell_K.value = f"={'F' + str(num_hrg)}-{'G' + str(num_hrg)}"
        cell_hasil.value = f"={cell_F.coordinate}*{avg}%"
        cell_hasil.number_format = 'Rp #,##0.00'
        ws.column_dimensions['G'].auto_size = True
        ws.column_dimensions['K'].auto_size = True
        ws['M1'].value = "Omset Bulanan ="
        ws['M2'].value = f"=SUM(E1:{'E' + str(len(ws['E']))})"
        ws['M3'].value = "Pengeluaran Bulanan = "
        ws['M4'].value = f"=SUM(J1:{'J' + str(len(ws['J']))})"
        ws['M5'].value = 'Sisa Saldo = '
        ws['M6'].value = f"=SUM(K1:{'K' + str(len(ws['K']))})"
        
    cell_F.number_format = 'Rp #,##0.00'
    cell_G.number_format = 'Rp #,##0.00'
    cell_K.number_format = 'Rp #,##0.00'
    ws['M2'].number_format = 'Rp #,##0.00'
    ws['M4'].number_format = 'Rp #,##0.00'
    ws['M6'].number_format = 'Rp #,##0.00'

    ws.column_dimensions['M'].auto_size = True
    wb.save('debit.xlsx')
    return cell_F.coordinate, cell_G.coordinate