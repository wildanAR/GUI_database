from openpyxl import Workbook, load_workbook
import datetime
import numpy as np


data_tgl = []
data_hrg = []
data_kdt = []

class ChartStat():
    def __init__(self):
        self.wb = load_workbook('audit/2023/kuartal_3.xlsx')

    def get_month(self):
        sheetName = self.wb.sheetnames
        return sheetName
        
    def cek(self):
        
        # for no,sheet in enumerate(wb.sheetnames):
        #     sheet = wb.worksheets[no]
            # print(sheet)
        sheet = wb.worksheets[2]
        for row in sheet:
            data = row[0].value
            if data is not None:
                data_tgl.append(data)
                data_hrg.append(row[4].value)
                data_kdt.append(row[9].value)
            for no,data in enumerate(data_tgl):
                if data == 'Tanggal':
                    del data_tgl[no]
                    del data_hrg[no]
                    del data_kdt[no]
        for no,data in enumerate(data_tgl):
            if type(data) is not str:
                data_tgl[no] = data.strftime('%m/%d/%Y')
        print(data_tgl)
        unique_list = []
        [unique_list.append(item) for item in data_tgl if item not in unique_list]
        pass
if __name__ == '__main__':
    sheet = statdebit().get_month()
    print(sheet)